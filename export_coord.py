"""
export_coord.py  —  PHL Catering Scheduler
COORD-format Excel export: mirrors the manual scheduling document layout.

Generates a two-sheet workbook:
  Sheet 1 — COORD        (domestic + precleared, sorted TRK → batch → stop)
  Sheet 2 — INT'L COORD  (true international only)

Call:
    from export_coord import generate_coord_excel
    xlsx_bytes = generate_coord_excel(result, day_label="Thursday", date_label="26 Mar 2026")

The result dict is the standard scheduler output:
    result['assignments'] — list of assignment dicts
    result['unassigned']  — list of unassigned flight dicts
"""

import io
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── TM-number → 2-digit TRK number (manual schedule convention) ──────────────
TM_TO_TRK = {
    'TM200':  1, 'TM201':  2, 'TM202':  3, 'TM203':  4, 'TM204':  5,
    'TM205':  6, 'TM206':  7, 'TM207':  8, 'TM208':  9,
    'TM100': 10, 'TM101': 11, 'TM102': 12, 'TM103': 13,
    'TM210': 15, 'TM211': 16, 'TM212': 17, 'TM214': 18,
    'TM215': 19, 'TM216': 20, 'TM217': 21, 'TM218': 22,
    'TM219': 23, 'TM220': 24, 'TM221': 25, 'TM222': 26,
    'TM223': 27, 'TM224': 28,
    'TM104': 29, 'TM105': 30, 'TM106': 31, 'TM107': 32, 'TM108': 33,
    'TM110': 34, 'TM111': 35,
}

def _trk(team_id: str) -> str:
    n = TM_TO_TRK.get(team_id)
    return str(n) if n else team_id  # fall back to TM### if not in map


# ── Colour palette ────────────────────────────────────────────────────────────
HDR_BG     = '1F3864'   # dark navy  — title bar
HDR_FG     = 'FFFFFF'
COL_BG     = '2E4D8A'   # mid navy   — column headers
COL_FG     = 'FFFFFF'
SEC_BG     = 'D9E1F2'   # pale blue  — TRK section divider
SEC_FG     = '1F3864'
OVN_BG     = 'E2EFDA'   # pale green — overnight teams (TM110/111)
OVN_FG     = '375623'
INTL_BG    = 'DDEBF7'   # sky blue   — international
INTL_FG    = '1F3864'
UNASSIGNED = 'FCE4D6'   # pale orange — unassigned / overflow

# Batch colours — alternating per team to distinguish consecutive batches
BATCH_COLS = [
    'FFFFFF',   # batch 1 — white
    'EEF4FB',   # batch 2 — very pale blue
    'F2F2F2',   # batch 3 — light grey
    'FFF9EC',   # batch 4 — pale cream
    'EBF7EE',   # batch 5 — pale mint
]

PRECLEARED = {
    'PUJ','CUN','AUA','STT','SXM','MBJ','NAS','PLS','EYW',
    'SDQ','SJU','BGI','GCM','MBJ','POS','SKB','VQS',
}

# ── Border helpers ────────────────────────────────────────────────────────────
_thin  = Side(style='thin',   color='BBBBBB')
_thick = Side(style='medium', color='888888')
_none  = Side(style=None)

def _border(top=None, bottom=None, left=None, right=None):
    return Border(
        top    = top    or _none,
        bottom = bottom or _none,
        left   = left   or _none,
        right  = right  or _none,
    )

_cell_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_group_top   = Border(left=_thin, right=_thin, top=_thick, bottom=_thin)
_group_bot   = Border(left=_thin, right=_thin, top=_thin,  bottom=_thick)
_group_mid   = Border(left=_thin, right=_thin, top=_thin,  bottom=_thin)
_group_solo  = Border(left=_thin, right=_thin, top=_thick, bottom=_thick)


# ── Cell writer ───────────────────────────────────────────────────────────────
def _w(cell, value, bg=None, fg='000000', bold=False, size=9,
       align='left', wrap=False, border=None):
    cell.value     = value
    cell.font      = Font(name='Calibri', size=size, bold=bold, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.border = border or _cell_border


# ── Column layout ─────────────────────────────────────────────────────────────
# Each tuple: (header_text, width, align)
COORD_COLS = [
    ('TRK',        5,  'center'),
    ('TEAM',       8,  'center'),
    ('OUT\nFLT',   7,  'center'),
    ('NOSE',       7,  'center'),
    ('GATE',       6,  'center'),
    ('EQP',        6,  'center'),
    ('FROM',       5,  'center'),
    ('IN',         7,  'center'),
    ('TO',         5,  'center'),
    ('OUT',        7,  'center'),
    ('TYPE',       12, 'left'),
    ('RUN\nSIZE',  5,  'center'),
    ('STOP',       4,  'center'),
    ('DISPATCH',   9,  'center'),
    ('SVC\nSTART', 9,  'center'),
    ('SVC\nEND',   9,  'center'),
    ('BATCH\n#',   5,  'center'),
    ('NOTES',      22, 'left'),
]


# ── Sheet builder ─────────────────────────────────────────────────────────────
def _build_sheet(ws, assignments, unassigned, title_text, show_overflow=True):
    ws.sheet_view.showGridLines = False

    # Column widths
    for col_idx, (_, w, _) in enumerate(COORD_COLS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row = 1

    # ── Title bar ─────────────────────────────────────────────────────────────
    n_cols = len(COORD_COLS)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    _w(ws.cell(row, 1), title_text, bg=HDR_BG, fg=HDR_FG, bold=True, size=12,
       align='center')
    ws.row_dimensions[row].height = 26
    row += 1

    # ── Column headers ────────────────────────────────────────────────────────
    for col_idx, (hdr, _, al) in enumerate(COORD_COLS, 1):
        _w(ws.cell(row, col_idx), hdr, bg=COL_BG, fg=COL_FG,
           bold=True, size=9, align='center', wrap=True)
    ws.row_dimensions[row].height = 28
    row += 1

    # ── Group assignments by team → dispatch batch ───────────────────────────
    # Key: (team_id, dispatch_min) → list of stop dicts sorted by stop_num
    by_team_batch = defaultdict(list)
    for a in assignments:
        team = a.get('team') or ''
        dmin = int(a.get('dispatch_min', 0))
        by_team_batch[(team, dmin)].append(a)

    # Sort batches within each team by dispatch_min, then by TRK number overall
    team_order = sorted(
        set(k[0] for k in by_team_batch),
        key=lambda t: (TM_TO_TRK.get(t, 99), t)
    )

    current_trk_bg   = None
    prev_team        = None
    batch_count      = {}   # team → running batch index (for colour cycling)

    for team_id in team_order:
        trk_num    = _trk(team_id)
        is_ovn     = team_id in ('TM110', 'TM111')
        team_bg_base = OVN_BG if is_ovn else 'FFFFFF'

        # ── TRK section divider row ───────────────────────────────────────────
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        section_txt = f"  TRK {trk_num}  —  {team_id}"
        _w(ws.cell(row, 1), section_txt, bg=SEC_BG, fg=SEC_FG, bold=True, size=9)
        ws.row_dimensions[row].height = 14
        row += 1

        # Get batches for this team, sorted by dispatch_min
        batches_for_team = sorted(
            [(dmin, stops) for (tid, dmin), stops in by_team_batch.items() if tid == team_id],
            key=lambda x: x[0]
        )
        batch_count[team_id] = 0

        for dmin, stops in batches_for_team:
            stops_sorted = sorted(stops, key=lambda s: int(s.get('stop_num', 1)))
            n_stops      = len(stops_sorted)
            b_idx        = batch_count[team_id] % len(BATCH_COLS)
            row_bg       = BATCH_COLS[b_idx]
            batch_count[team_id] += 1

            for s_idx, stop in enumerate(stops_sorted):
                is_first  = (s_idx == 0)
                is_last   = (s_idx == n_stops - 1)
                is_solo   = (n_stops == 1)

                if   is_solo:  bdr = _group_solo
                elif is_first: bdr = _group_top
                elif is_last:  bdr = _group_bot
                else:          bdr = _group_mid

                dest     = stop.get('dest', '')
                is_pre   = stop.get('is_precleared', False) or dest.upper() in PRECLEARED
                is_wb    = stop.get('is_wb', False)
                is_intl  = stop.get('is_intl', False) and not is_pre
                fg_col   = '000000'

                # Build note string
                notes = []
                if stop.get('is_short_turn'):  notes.append('SHORT TURN')
                if stop.get('is_tight_turn'):  notes.append('TIGHT TURN')
                if stop.get('is_ron'):         notes.append('RON')
                if is_pre:                     notes.append('PRECLEARED')
                if is_wb:                      notes.append('WIDEBODY')

                def _col(c): return c

                run_size_val = stop.get('run_size', 1)
                stop_num_val = stop.get('stop_num', 1)

                vals = [
                    trk_num if is_first else '',           # TRK
                    team_id if is_first else '',            # TEAM
                    stop.get('flight', ''),                 # OUT FLT
                    stop.get('nose', '') or '',             # NOSE
                    stop.get('gate', ''),                   # GATE
                    stop.get('equip', ''),                  # EQP
                    stop.get('inbound_origin', '') or '',   # FROM
                    stop.get('inbound_sta', '') or '',      # IN
                    dest,                                   # TO
                    stop.get('std', ''),                    # OUT
                    stop.get('type', ''),                   # TYPE
                    run_size_val,                           # RUN SIZE
                    stop_num_val,                           # STOP
                    stop.get('dispatch', ''),               # DISPATCH
                    stop.get('svc_start', ''),              # SVC START
                    stop.get('svc_end', ''),                # SVC END
                    batch_count[team_id],                   # BATCH #
                    ' | '.join(notes) if notes else '',     # NOTES
                ]

                for col_idx, (val, (_, _, al)) in enumerate(zip(vals, COORD_COLS), 1):
                    bold_flag = (col_idx <= 2 and is_first) or col_idx in (3, 9, 10)
                    _w(ws.cell(row, col_idx), val,
                       bg=row_bg, fg=fg_col,
                       bold=bold_flag, size=9,
                       align=COORD_COLS[col_idx - 1][2],
                       border=bdr)

                ws.row_dimensions[row].height = 15
                row += 1

        # Spacer between teams
        ws.row_dimensions[row].height = 4
        row += 1

    # ── Unassigned / overflow section ─────────────────────────────────────────
    if show_overflow and unassigned:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        _w(ws.cell(row, 1),
           f'  UNASSIGNED / OVERFLOW — {len(unassigned)} flights  '
           '(manual FH coverage or schedule gap)',
           bg='CC0000', fg='FFFFFF', bold=True, size=9)
        ws.row_dimensions[row].height = 16
        row += 1

        for u in sorted(unassigned, key=lambda x: int(x.get('std_min', 0))):
            dest = u.get('dest', '')
            vals = [
                'FH', '', u.get('flight', ''), u.get('nose', '') or '',
                u.get('gate', ''), u.get('equip', ''),
                '', '', dest, u.get('std', ''),
                u.get('type', ''), 1, 1, '', '', '',
                0, u.get('reason', 'UNASSIGNED'),
            ]
            for col_idx, (val, (_, _, al)) in enumerate(zip(vals, COORD_COLS), 1):
                _w(ws.cell(row, col_idx), val, bg=UNASSIGNED, size=9,
                   align=COORD_COLS[col_idx - 1][2])
            ws.row_dimensions[row].height = 15
            row += 1

    # Freeze header rows
    ws.freeze_panes = ws.cell(3, 1)


# ── Public entry point ────────────────────────────────────────────────────────
def generate_coord_excel(result: dict,
                         day_label: str = '',
                         date_label: str = '') -> bytes:
    """
    Generate the COORD + INT'L COORD Excel workbook.

    Parameters
    ----------
    result     : scheduler result dict (assignments + unassigned)
    day_label  : e.g. "Thursday"
    date_label : e.g. "26 Mar 2026"

    Returns
    -------
    bytes  — raw xlsx bytes, ready to return from a Flask Response.
    """
    assignments = result.get('assignments', []) or []
    unassigned  = result.get('unassigned',  []) or []

    label = f"{day_label}  {date_label}".strip()

    # ── Split into domestic and international ─────────────────────────────────
    dom_assigns  = []   # domestic + precleared
    intl_assigns = []   # true international only

    for a in assignments:
        dest  = (a.get('dest') or '').upper()
        is_pc = a.get('is_precleared', False) or dest in PRECLEARED
        is_i  = a.get('is_intl', False)
        if is_i and not is_pc:
            intl_assigns.append(a)
        else:
            dom_assigns.append(a)

    dom_unassigned  = [u for u in unassigned
                       if not (u.get('is_intl') and not u.get('is_precleared', False)
                               and (u.get('dest') or '').upper() not in PRECLEARED)]
    intl_unassigned = [u for u in unassigned
                       if u.get('is_intl') and not u.get('is_precleared', False)
                       and (u.get('dest') or '').upper() not in PRECLEARED]

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1 — COORD (domestic)
    ws_coord = wb.create_sheet('COORD')
    _build_sheet(
        ws_coord,
        dom_assigns,
        dom_unassigned,
        title_text=f'American Airlines Catering  ·  COORD  ·  {label}',
        show_overflow=True,
    )

    # Sheet 2 — INT'L COORD
    ws_intl = wb.create_sheet("INT'L COORD")
    _build_sheet(
        ws_intl,
        intl_assigns,
        intl_unassigned,
        title_text=f"American Airlines Catering  ·  INT'L COORD  ·  {label}",
        show_overflow=False,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Health Check tab ──────────────────────────────────────────────────────────
def _build_health_sheet(ws, hc: dict, label: str) -> None:
    """Add a Schedule Health Check sheet to an existing workbook sheet."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 38

    SEV_BG  = {'CRITICAL':'FFC7CE','HIGH':'FFEB9C','MEDIUM':'DDEBF7','INFO':'E2EFDA'}
    SEV_FG  = {'CRITICAL':'9C0006','HIGH':'9C6500','MEDIUM':'1F3864','INFO':'375623'}
    PASS_BG = 'E2EFDA'; FAIL_BG = 'FFC7CE'

    thin = Side(style='thin', color='CCCCCC')
    def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)
    def w(cell, val, bg=None, fg='000000', bold=False, size=9, align='left', wrap=False):
        cell.value = val
        cell.font  = Font(name='Calibri', size=size, bold=bold, color=fg)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if bg: cell.fill = PatternFill('solid', fgColor=bg)
        cell.border = bdr()

    r = 1
    n_cols = 5
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    w(ws.cell(r, 1), f'Schedule Health Check  ·  {label}',
      bg='1F3864', fg='FFFFFF', bold=True, size=13, align='center')
    ws.row_dimensions[r].height = 28; r += 1

    passed = hc.get('passed', True)
    summary = hc.get('summary', '')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    w(ws.cell(r, 1), summary, bg=PASS_BG if passed else FAIL_BG,
      bold=True, size=11, align='center')
    ws.row_dimensions[r].height = 22; r += 2

    for col, txt in enumerate(['','Severity','Code','Message','Detail / Flights'], 1):
        c = ws.cell(r, col)
        w(c, txt, bg='2E4D8A', fg='FFFFFF', bold=True, size=9, align='center')
    ws.row_dimensions[r].height = 20; r += 1

    order = [('CRITICAL', hc.get('critical',[])),
             ('HIGH',     hc.get('high',[])),
             ('MEDIUM',   hc.get('medium',[])),
             ('INFO',     hc.get('info',[]))]

    row_idx = 0
    for sev, items in order:
        if not items:
            continue
        bg  = SEV_BG[sev]
        fg  = SEV_FG[sev]
        alt = 'FFFFFF' if sev == 'INFO' else bg
        for item in items:
            row_bg = bg if row_idx % 2 == 0 else alt
            flights_str = ', '.join(str(f) for f in (item.get('flights') or [])[:5])
            detail = item.get('detail','')
            if flights_str:
                detail = f"{detail}  Flights: {flights_str}"
            w(ws.cell(r, 1), '', bg=row_bg)
            w(ws.cell(r, 2), sev, bg=bg, fg=fg, bold=True, size=9, align='center')
            w(ws.cell(r, 3), item.get('code',''), bg=row_bg, size=8)
            w(ws.cell(r, 4), item.get('message',''), bg=row_bg, bold=(sev=='CRITICAL'), size=9)
            w(ws.cell(r, 5), detail, bg=row_bg, size=8, wrap=True)
            ws.row_dimensions[r].height = 22
            r += 1
            row_idx += 1

    if not any(items for _, items in order):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        w(ws.cell(r, 1), '✓ No issues found.', bg=PASS_BG, bold=True, align='center')
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = ws.cell(4, 1)


def generate_coord_excel(result: dict,
                         day_label: str = '',
                         date_label: str = '') -> bytes:
    """
    Generate the COORD + INT'L COORD + Health Check Excel workbook.

    Parameters
    ----------
    result     : scheduler result dict (assignments + unassigned)
    day_label  : e.g. "Thursday"
    date_label : e.g. "26 Mar 2026"

    Returns
    -------
    bytes  — raw xlsx bytes, ready to return from a Flask Response.
    """
    assignments = result.get('assignments', []) or []
    unassigned  = result.get('unassigned',  []) or []

    label = f"{day_label}  {date_label}".strip()

    # ── Split domestic vs international ──────────────────────────────────────
    dom_assigns  = []
    intl_assigns = []
    for a in assignments:
        dest  = (a.get('dest') or '').upper()
        is_pc = a.get('is_precleared', False) or dest in PRECLEARED
        is_i  = a.get('is_intl', False)
        if is_i and not is_pc:
            intl_assigns.append(a)
        else:
            dom_assigns.append(a)

    dom_unassigned  = [u for u in unassigned
                       if not (u.get('is_intl') and not u.get('is_precleared', False)
                               and (u.get('dest') or '').upper() not in PRECLEARED)]
    intl_unassigned = [u for u in unassigned
                       if u.get('is_intl') and not u.get('is_precleared', False)
                       and (u.get('dest') or '').upper() not in PRECLEARED]

    wb = Workbook()
    wb.remove(wb.active)

    ws_coord = wb.create_sheet('COORD')
    _build_sheet(ws_coord, dom_assigns, dom_unassigned,
                 title_text=f'American Airlines Catering  ·  COORD  ·  {label}',
                 show_overflow=True)

    ws_intl = wb.create_sheet("INT'L COORD")
    _build_sheet(ws_intl, intl_assigns, intl_unassigned,
                 title_text=f"American Airlines Catering  ·  INT'L COORD  ·  {label}",
                 show_overflow=False)

    # ── Health Check tab ─────────────────────────────────────────────────────
    try:
        from scheduler_engine import health_check
        hc = health_check(result)
    except ImportError:
        hc = {'passed': True, 'critical': [], 'high': [], 'medium': [], 'info': [],
              'summary': '(health_check not available — update scheduler_engine.py)'}

    ws_hc = wb.create_sheet('Health Check')
    _build_health_sheet(ws_hc, hc, label)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
