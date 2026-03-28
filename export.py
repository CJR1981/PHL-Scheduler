"""
PHL Catering Scheduler — Excel Export v3 (Clean Light Theme)
9-tab workbook. Colors reserved for tight turns + internationals only.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, datetime

# ── Clean palette — neutral base, color only where it matters ─────────────
WHITE      = 'FFFFFF'
GRAY_BG    = 'F7F8FA'
GRAY_LIGHT = 'F0F1F4'
HDR_BG     = '1F2937'
HDR_FG     = 'FFFFFF'
TEXT_DARK  = '1F2937'
TEXT_MED   = '4B5563'
TEXT_LIGHT = '9CA3AF'
BORDER_CLR = 'E5E7EB'
INTL_BG    = 'EDE9FE';  INTL_TEXT  = '5B21B6'
ST_BG      = 'FEF2F2';  ST_TEXT    = 'B91C1C'
UNASGN_BG  = 'FEF2F2';  UNASGN_TEXT= 'DC2626'
OK_TEXT    = '059669';  WARN_TEXT  = 'D97706';  RED_TEXT = 'DC2626'
TITLE_BG   = '1E3A5F';  TITLE_FG   = 'FFFFFF'

def _fill(c): return PatternFill('solid', fgColor=c)
def _font(bold=False, color=TEXT_DARK, size=10): return Font(bold=bold, color=color, size=size, name='Calibri')
def _bdr(): return Border(bottom=Side(style='thin', color=BORDER_CLR))
def _tbdr(): return Border(bottom=Side(style='thin',color=BORDER_CLR),top=Side(style='thin',color=BORDER_CLR),left=Side(style='thin',color=BORDER_CLR),right=Side(style='thin',color=BORDER_CLR))
def _ctr(): return Alignment(horizontal='center', vertical='center')
def _lft(): return Alignment(horizontal='left', vertical='center')

def _hdr(ws, col, row, text, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(bold=True, color=HDR_FG, size=9); c.fill = _fill(HDR_BG)
    c.alignment = _ctr(); c.border = _tbdr()
    if width: ws.column_dimensions[get_column_letter(col)].width = width

def _title(ws, text, ncols, row=1):
    ws.row_dimensions[row].height = 28
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(bold=True, color=TITLE_FG, size=12); c.fill = _fill(TITLE_BG)
    c.alignment = _lft()
    for col in range(2, ncols+1): ws.cell(row=row, column=col).fill = _fill(TITLE_BG)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

def _meta(ws, stats, ncols, row=2):
    ws.row_dimensions[row].height = 18
    d = datetime.datetime.now().strftime('%A %d %b %Y  %H:%M')
    txt = f"PHL AA Catering  ·  {d}  ·  Coverage {stats.get('coverage_pct','?')}%  ·  {stats.get('assigned','?')}/{stats.get('total','?')} flights"
    c = ws.cell(row=row, column=1, value=txt)
    c.font = _font(color=TEXT_LIGHT, size=9); c.fill = _fill(GRAY_LIGHT)
    for col in range(2, ncols+1): ws.cell(row=row, column=col).fill = _fill(GRAY_LIGHT)

def _dc(ws, r, col, val, bg=WHITE, color=TEXT_DARK, bold=False, align='center'):
    c = ws.cell(row=r, column=col, value=val)
    c.fill = _fill(bg); c.font = _font(bold=bold, color=color)
    c.alignment = _ctr() if align=='center' else _lft(); c.border = _bdr()


# ═══════════════════════════════════════════════════════════════════════════
# TAB 1 — Full Schedule
# ═══════════════════════════════════════════════════════════════════════════
def _tab_schedule(wb, result):
    ws = wb.create_sheet('Full Schedule'); ws.sheet_view.showGridLines = False
    items = sorted(result['assignments'] + result['unassigned'], key=lambda x: x.get('std_min',0))

    cols = ['Flight','Dest','STD','Type','Gate','Equip','Nose','Pax','Inbound','From','Arr','Gnd Time','Team','Dispatch','Svc Start','Svc End']
    widths = [8,6,6,9,6,6,6,5,8,6,6,8,8,8,8,8]
    _title(ws, 'Full Flight Schedule', len(cols))
    _meta(ws, result['stats'], len(cols))
    ws.row_dimensions[3].height = 4
    for i,(h,w) in enumerate(zip(cols,widths),1): _hdr(ws,i,4,h,w)

    for idx, f in enumerate(items):
        row = idx + 5; ws.row_dimensions[row].height = 17
        ua = not f.get('team')
        intl = f.get('is_intl') and not f.get('is_precleared')
        st = f.get('is_short_turn') or f.get('is_tight_turn')
        if ua: bg, txt = UNASGN_BG, UNASGN_TEXT
        elif st: bg, txt = ST_BG, ST_TEXT
        elif intl: bg, txt = INTL_BG, INTL_TEXT
        else: bg, txt = (WHITE if idx%2==0 else GRAY_BG), TEXT_DARK

        tp = ('INTL' if intl else 'PRECLRD' if f.get('is_precleared') else 'DOM')
        if f.get('is_ron'): tp += ' RON'
        if st: tp += ' ST'
        if f.get('is_wb'): tp += ' WB'
        if ua: tp += ' ✗'

        inb = f'AA{f["inbound_flight"]}' if f.get('inbound_flight') else ('RON' if f.get('is_ron') else '')
        vals = [f.get('flight',''),f.get('dest',''),f.get('std',''),tp,f.get('gate',''),f.get('equip',''),
                f.get('nose',''),f.get('pax',''),inb,f.get('inbound_origin',''),f.get('inbound_sta',''),
                f.get('ground_time',''),f.get('team','UNASSIGNED'),f.get('dispatch',''),f.get('svc_start',''),f.get('svc_end','')]
        for col,val in enumerate(vals,1):
            fc = txt; b = col==1
            if col==13 and ua: fc=RED_TEXT; b=True
            if col==4 and (intl or st): b=True
            _dc(ws,row,col,val,bg=bg,color=fc,bold=b, align='left' if col==4 else 'center')

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{get_column_letter(len(cols))}4'


# ═══════════════════════════════════════════════════════════════════════════
# TAB 2 — Team Assignments
# ═══════════════════════════════════════════════════════════════════════════
def _tab_teams(wb, result):
    ws = wb.create_sheet('Team Assignments'); ws.sheet_view.showGridLines = False
    teams = sorted(result.get('team_summary',[]), key=lambda t: t.get('shift_start',0))
    cols = ['Team','Type','Shift','Cap','Flights','% Cap','First','Last','Flight List']
    widths = [8,6,12,5,7,7,7,7,60]
    _title(ws,'Team Assignments Summary',len(cols))
    _meta(ws,result['stats'],len(cols))
    ws.row_dimensions[3].height = 4
    for i,(h,w) in enumerate(zip(cols,widths),1): _hdr(ws,i,4,h,w)

    for idx, t in enumerate(teams):
        row = idx+5; ws.row_dimensions[row].height = 18
        fc = t.get('flight_count', t.get('ops',0)); cap = t.get('cap',5)
        pct = round(100*fc/cap) if cap else 0
        bg = WHITE if idx%2==0 else GRAY_BG
        ops = {o['flight']:o for o in t.get('operations',[])}
        ann = []
        for fn in t.get('flights',[]):
            op = ops.get(fn,{})
            tag = ' [I]' if (op.get('is_intl') and not op.get('is_precleared')) else ' [ST]' if op.get('is_short_turn') else ''
            ann.append(fn+tag)
        vals = [t['team_id'],t.get('team_type','FT'),t.get('shift',''),cap,fc,f"{pct}%",t.get('first_std',''),t.get('last_std',''),'   '.join(ann)]
        for col,val in enumerate(vals,1):
            color = TEXT_DARK; b = col==1
            if col==6 and pct>=100: color=WARN_TEXT
            _dc(ws,row,col,val,bg=bg,color=color,bold=b,align='left' if col==9 else 'center')
    ws.freeze_panes = 'A5'


# ═══════════════════════════════════════════════════════════════════════════
# TAB 3 — Short Turns
# ═══════════════════════════════════════════════════════════════════════════
def _tab_short_turns(wb, result):
    ws = wb.create_sheet('Short Turns'); ws.sheet_view.showGridLines = False
    items = sorted([f for f in result['assignments'] if f.get('is_short_turn')], key=lambda x:x.get('std_min',0))
    cols = ['Flight','Dest','Equip','Arr','STD','Gnd Time','Team','Dispatch','Svc Start','Svc End','Buf','Note']
    widths = [8,6,7,7,7,9,8,8,8,8,7,28]
    _title(ws, f'Short-Turn Flights  ({len(items)})', len(cols))
    _meta(ws, result['stats'], len(cols))
    ws.row_dimensions[3].height = 16
    c = ws.cell(row=3, column=1, value='These flights use relaxed finish buffer. Coordinate with gate teams.')
    c.font = _font(color=WARN_TEXT, size=9)
    for i,(h,w) in enumerate(zip(cols,widths),1): _hdr(ws,i,4,h,w)

    for idx, f in enumerate(items):
        row = idx+5; ws.row_dimensions[row].height = 18
        tight = f.get('is_tight_turn')
        vals = [f.get('flight',''),f.get('dest',''),f.get('equip',''),
                f.get('arrival_time_str',f.get('inbound_sta','—')),f.get('std',''),
                f.get('ground_time_str',f.get('ground_time','—')),f.get('team',''),
                f.get('dispatch',''),f.get('svc_start',''),f.get('svc_end',''),
                'TIGHT' if tight else '25 min',
                'TIGHT TURN' if tight else 'Short turn — relaxed buffer']
        for col,val in enumerate(vals,1):
            b = col==12 or (col==11 and tight)
            _dc(ws,row,col,val,bg=ST_BG,color=ST_TEXT,bold=b, align='left' if col==12 else 'center')

    if not items:
        ws.cell(row=5,column=1,value='No short-turn flights.').font = _font(color=OK_TEXT)
    ws.freeze_panes = 'A5'


# ═══════════════════════════════════════════════════════════════════════════
# TAB 4 — International
# ═══════════════════════════════════════════════════════════════════════════
def _tab_intl(wb, result):
    ws = wb.create_sheet('International'); ws.sheet_view.showGridLines = False
    items = sorted([f for f in result['assignments'] if f.get('is_intl') and not f.get('is_precleared')], key=lambda x:x.get('std_min',0))
    cols = ['Flight','Dest','Equip','WB/NB','STD','Gate','Team','Dock Load','Dispatch','Svc Start','Svc End','Bank']
    widths = [8,6,7,6,7,6,8,8,8,8,8,10]
    _title(ws, f'International Flights  ({len(items)})', len(cols))
    _meta(ws, result['stats'], len(cols))
    for i,(h,w) in enumerate(zip(cols,widths),1): _hdr(ws,i,3,h,w)

    for idx, f in enumerate(items):
        row = idx+4; ws.row_dimensions[row].height = 18
        vl = 'Evening WB' if f.get('is_evening_wb') else 'Morning NB' if f.get('is_morning_intl') else 'Afternoon'
        vals = [f.get('flight',''),f.get('dest',''),f.get('equip',''),'WB' if f.get('is_wb') else 'NB',
                f.get('std',''),f.get('gate',''),f.get('team',''),f.get('dock_load',''),
                f.get('dispatch',''),f.get('svc_start',''),f.get('svc_end',''),vl]
        for col,val in enumerate(vals,1):
            b = col==1 or (col==4 and f.get('is_wb'))
            _dc(ws,row,col,val,bg=INTL_BG,color=INTL_TEXT,bold=b)
    ws.freeze_panes = 'A4'


# ═══════════════════════════════════════════════════════════════════════════
# TAB 5 — Overnight & RON
# ═══════════════════════════════════════════════════════════════════════════
def _tab_overnight(wb, result):
    ws = wb.create_sheet('Overnight & RON'); ws.sheet_view.showGridLines = False
    items = sorted([f for f in result['assignments'] if f.get('is_ron')], key=lambda x:x.get('std_min',0))
    cols = ['Flight','Dest','STD','Equip','Gate','Team','Dispatch','Svc Start','Svc End','Note']
    widths = [8,6,7,7,6,8,8,8,8,28]
    _title(ws, f'Overnight & RON  ({len(items)} RON flights)', len(cols))
    _meta(ws, result['stats'], len(cols))
    ws.row_dimensions[3].height = 16
    c = ws.cell(row=3,column=1,value='RON aircraft already at gate. Overnight teams dispatch from 23:15.')
    c.font = _font(color=TEXT_LIGHT, size=9)
    for i,(h,w) in enumerate(zip(cols,widths),1): _hdr(ws,i,4,h,w)

    for idx, f in enumerate(items):
        row = idx+5; ws.row_dimensions[row].height = 18
        ovn = f.get('team','') in ('TM110','TM111')
        bg = WHITE if idx%2==0 else GRAY_BG
        vals = [f.get('flight',''),f.get('dest',''),f.get('std',''),f.get('equip',''),f.get('gate',''),
                f.get('team',''),f.get('dispatch',''),f.get('svc_start',''),f.get('svc_end',''),
                'Overnight team' if ovn else 'RON — early team']
        for col,val in enumerate(vals,1):
            _dc(ws,row,col,val,bg=bg,color=TEXT_DARK,align='left' if col==10 else 'center')
    ws.freeze_panes = 'A5'


# ═══════════════════════════════════════════════════════════════════════════
# TAB 6 — Unassigned
# ═══════════════════════════════════════════════════════════════════════════
def _tab_unassigned(wb, result):
    ws = wb.create_sheet('Unassigned'); ws.sheet_view.showGridLines = False
    ua = sorted(result.get('unassigned',[]), key=lambda x:x.get('std_min',0))
    RL = {'NO_TRUCK_AVAILABLE':'All trucks at capacity','OUTSIDE_SHIFT_WINDOW':'No team covers this time',
          'NO_TEAM_AVAILABLE':'All teams at cap','NO_TEAM_CAPACITY':'All teams at cap',
          'INSUFFICIENT_GROUND_TIME':'Ground time too short'}
    AL = {'NO_TRUCK_AVAILABLE':'Manual reassign or stagger flights','OUTSIDE_SHIFT_WINDOW':'Extend shift or add coverage',
          'NO_TEAM_AVAILABLE':'Reassign lower-priority flight','NO_TEAM_CAPACITY':'Reassign lower-priority flight',
          'INSUFFICIENT_GROUND_TIME':'Short-turn protocol with gate'}
    cols = ['Flight','Dest','STD','Type','Gate','Equip','Reason','Action Required']
    widths = [8,6,7,8,6,7,28,35]
    _title(ws, f'Unassigned Flights  ({len(ua)})', len(cols))
    _meta(ws, result['stats'], len(cols))
    for i,(h,w) in enumerate(zip(cols,widths),1): _hdr(ws,i,3,h,w)
    if not ua:
        ws.cell(row=4,column=1,value='All flights assigned — 100% coverage.').font = _font(color=OK_TEXT,bold=True)
        return
    for idx, f in enumerate(ua):
        row = idx+4; r = f.get('reason','?')
        tp = ('INTL' if f.get('is_intl') else 'DOM') + (' RON' if f.get('is_ron') else '')
        vals = [f.get('flight',''),f.get('dest',''),f.get('std',''),tp,f.get('gate',''),f.get('equip',''),
                RL.get(r,r),AL.get(r,'Coordinator action required')]
        for col,val in enumerate(vals,1):
            _dc(ws,row,col,val,bg=UNASGN_BG,color=RED_TEXT if col<=6 else WARN_TEXT,bold=col==1,align='left' if col>=7 else 'center')
    ws.freeze_panes = 'A4'


# ═══════════════════════════════════════════════════════════════════════════
# TAB 7 — Truck Utilisation
# ═══════════════════════════════════════════════════════════════════════════
def _tab_trucks(wb, result):
    ws = wb.create_sheet('Truck Utilisation'); ws.sheet_view.showGridLines = False
    TC = 25; slots = {}
    for a in result['assignments']:
        d = a.get('dispatch_min'); fr = a.get('team_free_min')
        if d is None or fr is None: continue
        for h in range(24):
            for half in range(2):
                ss = h*60+half*30
                if d < ss+30 and fr > ss:
                    slots.setdefault(h*2+half,[]).append(a['flight'])
    cols = ['Time','In Use','Capacity','Util %','Flights']
    widths = [7,8,8,8,55]
    _title(ws, f'Truck Utilisation (Pool = {TC})', len(cols))
    _meta(ws, result['stats'], len(cols))
    for i,(h,w) in enumerate(zip(cols,widths),1): _hdr(ws,i,3,h,w)
    row = 4
    for sk in sorted(slots.keys()):
        fl = slots[sk]; n = len(fl)
        if not n: continue
        h = sk//2; half = sk%2; pct = round(100*n/TC)
        bg = WHITE if row%2==0 else GRAY_BG
        pc = RED_TEXT if n>TC else WARN_TEXT if pct>=80 else OK_TEXT if pct>=40 else TEXT_LIGHT
        _dc(ws,row,1,f"{h:02d}:{('00' if half==0 else '30')}",bg=bg,color=TEXT_DARK,bold=True)
        _dc(ws,row,2,n,bg=bg,color=pc,bold=True)
        _dc(ws,row,3,TC,bg=bg,color=TEXT_LIGHT)
        _dc(ws,row,4,f"{pct}%",bg=bg,color=pc,bold=True)
        _dc(ws,row,5,' '.join(fl[:15])+(' …' if len(fl)>15 else ''),bg=bg,color=TEXT_MED,align='left')
        row += 1
    ws.freeze_panes = 'A4'


# ═══════════════════════════════════════════════════════════════════════════
# TAB 8 — Coverage Gaps
# ═══════════════════════════════════════════════════════════════════════════
def _tab_gaps(wb, result):
    ws = wb.create_sheet('Coverage Gaps'); ws.sheet_view.showGridLines = False
    from collections import defaultdict
    bh = defaultdict(lambda:{'assigned':0,'unassigned':0,'intl':0,'st':0,'ron':0})
    for a in result['assignments']:
        h = (a.get('std_min') or 0)//60; bh[h]['assigned']+=1
        if a.get('is_intl'): bh[h]['intl']+=1
        if a.get('is_short_turn'): bh[h]['st']+=1
        if a.get('is_ron'): bh[h]['ron']+=1
    for u in result.get('unassigned',[]):
        h = (u.get('std_min') or 0)//60; bh[h]['unassigned']+=1
    cols = ['Hour','Assigned','Unassigned','INTL','Short Turn','RON','Total','Coverage']
    widths = [7,9,10,7,10,6,7,9]
    _title(ws,'Coverage by Hour',len(cols))
    _meta(ws,result['stats'],len(cols))
    for i,(h,w) in enumerate(zip(cols,widths),1): _hdr(ws,i,3,h,w)
    for idx, hour in enumerate(sorted(bh.keys())):
        row = idx+4; d = bh[hour]; tot = d['assigned']+d['unassigned']
        pct = round(100*d['assigned']/tot) if tot else 100
        gap = d['unassigned']>0; bg = UNASGN_BG if gap else (WHITE if idx%2==0 else GRAY_BG)
        vals = [f"{hour:02d}:00",d['assigned'],d['unassigned'] or '',d['intl'] or '',d['st'] or '',d['ron'] or '',tot,f"{pct}%"]
        for col,val in enumerate(vals,1):
            c=TEXT_DARK; b=False
            if col==3 and d['unassigned']: c=RED_TEXT; b=True
            elif col==8: c=OK_TEXT if pct==100 else RED_TEXT; b=True
            _dc(ws,row,col,val,bg=bg,color=c,bold=b)
    ws.freeze_panes = 'A4'


# ═══════════════════════════════════════════════════════════════════════════
# TAB 9 — Change Log
# ═══════════════════════════════════════════════════════════════════════════
def _tab_changelog(wb, change_log):
    ws = wb.create_sheet('Change Log'); ws.sheet_view.showGridLines = False
    cols = ['Time','Operation','Detail','Affected','Result']; widths = [8,14,35,30,25]
    n = len(change_log) if change_log else 0
    _title(ws, f'Change Log  ({n} change{"s" if n!=1 else ""})', len(cols))
    for i,(h,w) in enumerate(zip(cols,widths),1): _hdr(ws,i,2,h,w)
    if not change_log:
        ws.cell(row=3,column=1,value='No live operations performed.').font = _font(color=TEXT_LIGHT); return
    for idx, e in enumerate(change_log):
        row = idx+3; op = e.get('type','?'); bg = WHITE if idx%2==0 else GRAY_BG
        if op=='sick_call':
            det = f"{e.get('team','?')} called out"; aff = ', '.join(f"{r['flight']}→{r['new_team']}" for r in (e.get('breakdown') or []))
            res = f"{e.get('reassigned',0)} reassigned"
        elif op=='delay':
            det = f"Flight {e.get('flight','?')} delayed {e.get('from','?')}→{e.get('to','?')}"; aff = e.get('team','?')
            res = 'Conflict' if e.get('conflict') else 'No conflict'
        else:
            det = f"Flight {e.get('flight','?')} reassigned"; aff = f"{e.get('from','?')} → {e.get('to','?')}"; res = 'Applied'
        for col,val in enumerate([e.get('time',''),op.replace('_',' ').title(),det,aff,res],1):
            _dc(ws,row,col,val,bg=bg,color=TEXT_DARK,align='left')
    ws.freeze_panes = 'A3'


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════
def generate_excel(result: dict, change_log: list = None) -> bytes:
    wb = Workbook(); wb.remove(wb.active)
    _tab_schedule(wb, result)
    _tab_teams(wb, result)
    _tab_short_turns(wb, result)
    _tab_intl(wb, result)
    _tab_overnight(wb, result)
    _tab_unassigned(wb, result)
    _tab_trucks(wb, result)
    _tab_gaps(wb, result)
    _tab_changelog(wb, change_log or [])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
